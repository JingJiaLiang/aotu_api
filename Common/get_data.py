# -*- coding: utf-8 -*-
# @Time : 2020/10/26 23:34
# @Author : Liangjiajing
# @FileName: get_data.py
# @Email : 1369462217@qq.com
from openpyxl import load_workbook
from Common.dir_config import *
class Data:
    Cookie = None

class GetId:
    file_name = os.path.join(testdatas_dir, 'do_excel.xlsx')
    wb = load_workbook(file_name)
    sheet = wb['init']
    id = int(sheet.cell(2,1).value)+1
print(GetId.id)