# -*- coding: utf-8 -*-
# @Time : 2021/10/15 21:19
# @Author : Liangjiajing
# @FileName: DoExcel.py
# @Email : 1369462217@qq.com
from openpyxl import load_workbook
from Common.dir_config import *
from Common.get_data import GetId

import os
class DoExcel:
    # 获取excel里的数据
    def do_excel(self,file_name,sheet_name):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        test_data = []
        for item in range(2,sheet.max_row+1):
            sun_data = {'case_id': sheet.cell(item, 1).value, 'case_name': sheet.cell(item, 2).value,
                        'url': sheet.cell(item, 3).value, 'method': sheet.cell(item, 4).value,
                        'data': sheet.cell(item, 5).value, 'msg': sheet.cell(item, 6).value}
            if sheet.cell(item,5).value.find('${id}') != -1:
                sun_data['data'] = sheet.cell(item,5).value.replace('${id}',str(getattr(GetId,('id'))))
            #  sun_data['data'] = sheet.cell(item,5).value.replace('${id}',str(getattr(Data,'id')))

            test_data.append(sun_data)
        return test_data
    # 返回result
    def write_back(self,file_name,sheet_name,i,value,TestResult=None):
        wb = load_workbook(file_name)
        sheet = wb[sheet_name]
        sheet.cell(i, 7).value = value
        sheet.cell(i, 8).value = TestResult   #写回pass/failed
        #保存文件
        wb.save(file_name)
    #获取id
    def get_id(self,value):
        file_name = os.path.join(testdatas_dir, 'do_excel.xlsx')
        wb = load_workbook(file_name)
        sheet = wb['init']
        sheet.cell(2,1).value = value
        wb.save(file_name)

if __name__ == '__main__':
    file_name = os.path.join(testdatas_dir, 'do_excel.xlsx')
    print(DoExcel().do_excel(file_name,'teacher'))
































